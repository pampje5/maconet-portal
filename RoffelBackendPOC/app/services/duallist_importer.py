import logging
import openpyxl
from sqlalchemy.orm import Session
from app.models.article import Article

log = logging.getLogger(__name__)

# ======================
# PRIJSFACTOREN
# ======================
PURCHASE_FACTOR = 0.66
BRUTO_FACTOR = 1.50
WVK_FACTOR = 1.25
EDMAC_FACTOR = 1.05


# ======================
# IMPORTER
# ======================
def import_duallist_from_excel(file_path: str, db: Session) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # ----------------------
    # Sheet bepalen
    # ----------------------
    if "Duallist" in wb.sheetnames:
        ws = wb["Duallist"]
        sheet_used = "Duallist"
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_used = wb.sheetnames[0]

    log.info("Duallist import gestart (sheet=%s)", sheet_used)

    created = 0
    updated = 0
    skipped = 0
    duplicates_in_file = 0

    seen_part_nos: set[str] = set()

    # ----------------------
    # Loop door Excel
    # ----------------------
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        part_no = row[0]
        description = row[1]
        list_price = row[2]

        # Basisvalidatie
        if not part_no or list_price is None:
            skipped += 1
            log.debug("Rij %s overgeslagen (lege part_no of prijs)", row_idx)
            continue

        part_no = str(part_no).strip()

        # 🔥 DEDUPLICATIE BINNEN HET BESTAND
        if part_no in seen_part_nos:
            duplicates_in_file += 1
            log.warning(
                "Dubbele artikelcode in bestand (%s) op rij %s",
                part_no,
                row_idx,
            )
            continue

        seen_part_nos.add(part_no)

        try:
            list_price = float(list_price)
        except (TypeError, ValueError):
            skipped += 1
            log.warning(
                "Ongeldige list_price voor %s op rij %s",
                part_no,
                row_idx,
            )
            continue

        # ----------------------
        # Prijzen berekenen
        # ----------------------
        price_purchase = list_price * PURCHASE_FACTOR
        price_bruto = list_price * BRUTO_FACTOR
        price_wvk = list_price * WVK_FACTOR
        price_edmac = list_price * EDMAC_FACTOR

        # ----------------------
        # UPSERT
        # ----------------------
        existing = (
            db.query(Article)
            .filter(Article.part_no == part_no)
            .first()
        )

        if existing:
            existing.description = description or ""
            existing.list_price = list_price
            existing.price_purchase = price_purchase
            existing.price_bruto = price_bruto
            existing.price_wvk = price_wvk
            existing.price_edmac = price_edmac
            updated += 1
        else:
            db.add(
                Article(
                    part_no=part_no,
                    description=description or "",
                    list_price=list_price,
                    price_purchase=price_purchase,
                    price_bruto=price_bruto,
                    price_wvk=price_wvk,
                    price_edmac=price_edmac,
                    active=True,
                )
            )
            created += 1

    # ----------------------
    # Commit
    # ----------------------
    db.commit()

    log.info(
        "Duallist import afgerond: created=%s updated=%s skipped=%s duplicates=%s",
        created,
        updated,
        skipped,
        duplicates_in_file,
    )

    return {
        "sheet_used": sheet_used,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "duplicates_in_file": duplicates_in_file,
        "total_processed": created + updated,
    }
